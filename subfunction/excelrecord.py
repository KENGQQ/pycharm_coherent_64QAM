from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import openpyxl
import datetime
import os

def open_excel(address):
    if not os.path.exists(address + 'Record.xlsx'):
        print('creat an excel file')
        wb = Workbook()
        ws = wb.create_sheet('trial_1')
        del wb['Sheet']

    else:
        wb = openpyxl.load_workbook(address + 'Record.xlsx')
        ws = wb.create_sheet('trial_' + str(len(wb.worksheets) + 1), 0)

    ws['A1'] = datetime.datetime.now()

    parameter = ['eyepos', 'CMA[mean, type, overhead, earlystop, step_adj]', \
                 'CMA1[taps, stepX, stepY, iter, cost]', 'CMA2[taps, stepX, stepY, iter, cost]', 'CMA3[taps, stepX, stepY, iter, cost]', \
                  'PLL BW', 'VV radius', 'X_corr ',
                 ' X[SNR,EVM,BERcount]', 'Y_corr ', 'Y[SNR,EVM,BERcount]', 'corr vector(X, Y)']

    ws.column_dimensions['A'].width = 20.0
    ws.column_dimensions['B'].width =  6.0
    ws.column_dimensions['C'].width = 42.0
    ws.column_dimensions['D'].width = 35.0
    ws.column_dimensions['E'].width = 35.0
    ws.column_dimensions['F'].width = 35.0
    ws.column_dimensions['G'].width = 10.0
    ws.column_dimensions['H'].width = 20.0
    ws.column_dimensions['I'].width = 30.0
    ws.column_dimensions['J'].width = 25.0
    ws.column_dimensions['K'].width = 30.0
    ws.column_dimensions['L'].width = 25.0
    ws.column_dimensions['M'].width = 27.0

    ws.row_dimensions[1].height = 30

    for i in range(len(parameter)):
        ws.cell(row=1, column=i + 2, value=parameter[i])

    wb.save(address + 'Record.xlsx')


def write_excel(address, parameter_record):
    wb = openpyxl.load_workbook(address + 'Record.xlsx')
    ws = wb.worksheets[0]
    maxrow = ws.max_row
    for i in range(len(parameter_record)):
        ws.cell(row=maxrow + 1, column=1, value=datetime.datetime.now())
        ws.cell(row=maxrow + 1, column=i + 2, value=parameter_record[i])

    fill = PatternFill("solid", fgColor="d9b3ff")
    if (parameter_record[7][8:12] == parameter_record[7][2:6] ):
        for cell in list(ws.rows)[maxrow][8 : 10]:
            cell.fill = fill
    #
    fill = PatternFill("solid", fgColor="88F1A1")
    if (parameter_record[9][8:12] == parameter_record[9][2:6] ):
        for cell in list(ws.rows)[maxrow][10 : 12]:
            cell.fill = fill

    wb.save(address + 'Record.xlsx')
